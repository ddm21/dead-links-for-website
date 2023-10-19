import requests
import openpyxl
from openpyxl.styles import Font
from bs4 import BeautifulSoup
from tqdm import tqdm

# Function to check if a link is alive
def is_link_alive(url):
    try:
        response = requests.head(url)
        return response.status_code == 200
    except requests.ConnectionError:
        return False

# Function to extract links from a page
def get_links_from_page(url):
    try:
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')
        links = []
        for link in soup.find_all('a'):
            href = link.get('href')
            if href and href.startswith('http'):
                links.append(href)
        return links
    except requests.ConnectionError:
        return []

# URL of the website you want to check
website_url = 'https://example.com'

# Create a new Excel workbook and add a worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Set the headers for the Excel file
worksheet['A1'] = 'Page URL'
worksheet['B1'] = 'Element'
worksheet['C1'] = 'Link'
worksheet['D1'] = 'Status'

# Style the header row
font = Font(bold=True)
for cell in worksheet['1:1']:
    cell.font = font

# Initialize row counter
row = 2

# Check links on the website
pages_to_check = [website_url]

# Use tqdm to create a progress bar
with tqdm(total=len(pages_to_check), unit='page') as page_pbar:
    for page_url in pages_to_check:
        links_to_check = get_links_from_page(page_url)
        with tqdm(total=len(links_to_check), unit='link', leave=False) as link_pbar:
            for link in links_to_check:
                if is_link_alive(link):
                    status = '✔️'
                else:
                    status = '❌'
                worksheet.cell(row=row, column=1, value=page_url)
                worksheet.cell(row=row, column=2, value='Link')
                worksheet.cell(row=row, column=3, value=link)
                worksheet.cell(row=row, column=4, value=status)
                row += 1
                link_pbar.update(1)
        page_pbar.update(1)

# Save the Excel file
workbook.save('dead_links_report.xlsx')

print("Dead links report generated in 'dead_links_report.xlsx'")
